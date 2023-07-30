import openpyxl
from datetime import datetime

def get_input():
    roll_number = input("Enter Roll Number: ")
    name = input("Enter Name: ")
    return roll_number, name

def display_success_message(roll_number, name):
    print("\nSuccessfully stored the following data:")
    print("Roll Number:", roll_number)
    print("Name:", name)

def read_existing_data(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        return []
    
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

def save_to_excel(file_name, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row_data in data:
        sheet.append(row_data)

    workbook.save(file_name)

if __name__ == "__main__":
    timestamp = datetime.now().strftime("%d%m%y")
    file_name = f"student_data_{timestamp}.xlsx"

    existing_data = read_existing_data(file_name)

    while True:
        roll_number, name = get_input()
        display_success_message(roll_number, name)

        # Append the new data to the existing_data
        existing_data.append([roll_number, name])

        save_to_excel(file_name, existing_data)

        another_entry = input("\nDo you want to enter another data (yes/no)? ").lower()
        if another_entry != "yes":
            print("Exiting program.")
            break
